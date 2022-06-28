from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color
from PyQt6 import QtCore
from PyQt6.QtCore import pyqtSlot, QObject, pyqtSignal
from datetime import datetime

from m_config.config import (
    conf,
    TIMELOG_SOURCE_PATH,
    EXPENSES_SOURCE_PATH,
)
from m_folder_handler.folder_handler import FolderHandler


DATE_CELL = "F2"
NAME_CELL = "E4"
JOB_TYTLE_CELL = "M4"
GREY_COLOR = "BB808080"

day_status = {
    0: {"col": "K", "value": "HO"},  # Home Office
    1: "",  # Irodai munkanap
    2: {"col": "I", "value": "B"},  # Beteg szabadság
    3: {"col": "I", "value": "Fsz"},  # Fizetett szabadság
    4: {"col": "I", "value": "Figt"},  # Fizetett igazolt távollét
    5: {"col": "I", "value": "Áll"},  # Elrendelt fizetett állásidő
    6: {"col": "H", "value": 1},  # Külföldi munka
    7: {"col": "J", "value": "Fnsz"},  # Nem fizetett szabadság
    8: {"col": "J", "value": "Nigt"},  # Nem fizetett igazolt távollét
    9: {"col": "J", "value": "H"},  # Igazolatlan hiányzás
    10: {"col": "I", "value": "Fü"},  # Fizetett ünnep
    11: {""},  # Hetvege
}

temp_dir_name = ""


class TimeDocGenerator(QObject):
    data_updated_signal = pyqtSignal()

    def __init__(self, parent=None) -> None:
        super().__init__(parent=parent)
        self.wb = Workbook()
        self.ws = None
        self.num_of_wds = 0
    
    def set_num_of_wds(self, wds:int) -> None:
        self.num_of_wds = wds

    def load(self, path: str = None):
        try:
            self.wb = load_workbook(TIMELOG_SOURCE_PATH)
        except FileNotFoundError as e:
            print("Error: ", e)
            return

        self.ws = self.wb["YYYY MM"]
        print("Load success")

    def delete_extra_days(self):
        print(f"Call extra days delete, num of days: {conf.num_of_days}")
        if conf.num_of_days == 31:
            return
        num_to_del = 31 - conf.num_of_days
        end_of_days_cell = 40 - num_to_del
        print(f"Delete rows: {num_to_del}")
        self.ws.delete_rows(41 - num_to_del, num_to_del)
        self.ws[
            f"E{41-num_to_del}"
        ] = f"=SUM(E10:E{end_of_days_cell})+SUM(F10:F{end_of_days_cell})"
        self.ws[f"E{42-num_to_del}"] = f"=SUM(H10:H{end_of_days_cell})"
        self.ws[
            f"B{43-num_to_del}"
        ] = f'=SUBSTITUTE(SUBSTITUTE(SUBSTITUTE("Össz.kötelező munkanap: $1 nap, $2 fiz.ü.nap. (telj.munkaidő össz.óra: $3 óra)", "$1",  {self.num_of_wds}), "$2", E{50-num_to_del}), "$3", {self.num_of_wds} * 8)'
        self.ws[f"E{50-num_to_del}"] = f'=COUNTIF(I10:I{end_of_days_cell}, "Fü")'
        self.ws[f"E{51-num_to_del}"] = f'=COUNTIF(I10:I{end_of_days_cell}, "Fsz")'
        self.ws[f"E{52-num_to_del}"] = f'=COUNTIF(I10:I{end_of_days_cell}, "Figt")'
        self.ws[f"E{53-num_to_del}"] = f'=COUNTIF(K10:K{end_of_days_cell}, "HO")'
        self.ws[f"E{54-num_to_del}"] = f'=COUNTIF(I10:I{end_of_days_cell}, "Áll")'
        self.ws[f"E{55-num_to_del}"] = f'=COUNTIF(I10:I{end_of_days_cell}, "B")'
        self.ws.merge_cells(f"B{43-num_to_del}:N{43-num_to_del}")

    def set_user_data(self):
        self.ws.title = conf.now.strftime("%Y %b")
        self.ws[NAME_CELL] = f"{conf._last_name} {conf._first_name}"
        self.ws[DATE_CELL] = conf.now.strftime("%Y %b")
        self.ws[JOB_TYTLE_CELL] = f"{conf._role}"

    def set_days(self, days: list = None):
        for day in days:
            row = int(day["idx"] + 9)
            status_idx = day["status"]

            # First clear the previusly set data from rows
            for col in range(8, 12):
                self.ws.cell(row, col).value = ""
            if status_idx == 1:
                continue
            
            if status_idx > 1 and status_idx < 10 and status_idx != 6:
                self.del_data_row(row)

            if status_idx == 11:
                self.del_data_and_color_row(row)
                # print("Color to grey, remove fields, its weekend")
                continue
            if status_idx == 10:
                self.del_data_and_color_row(row)
                # print("Color to grey, remove fields, its a payd holiday")

            status_cell = day_status[status_idx]["col"] + str(row)
            value = day_status[status_idx]["value"]
            self.ws[status_cell] = value

        self.data_updated_signal.emit()
    
    def del_data_row(self, row):
        for col in range(3, 7):
            self.ws.cell(row, col).value = ""

    def del_data_and_color_row(self, row):
        my_gray = Color(rgb="00808080")
        my_fill = PatternFill(patternType="solid", fgColor=my_gray)
        for col in range(3, 7):
            self.ws.cell(row, col).value = ""
        for col in range(2, 15):
            self.ws.cell(row, col).fill = my_fill

    # Loads the Excel templte, sets the user data.
    def setup(self, wds:int):
        self.load()
        self.set_num_of_wds(wds)
        self.delete_extra_days()
        self.set_user_data()  # Todo move this to the end.

    @QtCore.pyqtSlot(list)
    def update_days(self, days: list):
        self.set_days(days)

    @QtCore.pyqtSlot()
    def save(self):
        FolderHandler.clear_create()
        self.set_user_data()
        self.wb.save(conf.get_timelog_path())

    @QtCore.pyqtSlot()
    def save(self):
        FolderHandler.clear_create()
        self.set_user_data()
        self.wb.save(conf.get_timelog_path())


class ExpensesDocGenerator(QObject):
    ready_to_save_signal = pyqtSignal()
    saved_signal = pyqtSignal()
    purpose_cell = "C3"
    name_cell = "C6"
    position_cell = "F6"
    from_cell = "M5"
    to_cell = "M6"
    exp_start_r = 10
    exp_date_c = "B"
    exp_desc_c = "C"
    exp_hotel_c = "D"
    exp_trans_c = "E"
    exp_fuel_c = "F"
    exp_allow_c = "G"
    exp_phone_c = "H"
    exp_ent_c = "I"
    exp_other_c = "J"
    exp_curr_c = "K"
    exp_exc_rate_c = "M"

    def __init__(self, parent=None) -> None:
        super().__init__(parent=parent)
        self.wb = Workbook()
        self.ws = None
        self.files_to_copy = []

    def load(self):
        try:
            self.wb = load_workbook(EXPENSES_SOURCE_PATH)
        except FileNotFoundError as e:
            print("Error: ", e)
            return

        self.ws = self.wb["expense"]
        print("Load success")

    def set_user_data(self):
        self.ws[self.name_cell] = f"{conf._first_name} {conf._last_name}"
        self.ws[self.position_cell] = f"{conf._role}"
        self.ws[self.purpose_cell] = "Budapest, {}".format(conf.now.strftime("%B %Y"))
        self.ws[self.from_cell] = datetime(conf.now.year, conf.now.month, 1).strftime(
            "%Y-%m-%d"
        )
        self.ws[self.to_cell] = datetime(
            conf.now.year, conf.now.month, conf.num_of_days
        ).strftime("%Y-%m-%d")

    # Loads the Excel templte, sets the user data.
    def setup(self):
        self.load()
        self.set_user_data()  # Todo move this to the end.

    @pyqtSlot(list)
    def receive_expense_data_slot(self, expenses: list):
        exp_type_table = {
            "Hotel": self.exp_hotel_c,
            "Transport": self.exp_trans_c,
            "Fuel": self.exp_fuel_c,
            "Allowance": self.exp_allow_c,
            "Phone": self.exp_phone_c,
            "Entertainment": self.exp_ent_c,
            "Other": self.exp_other_c,
        }

        self.files_to_copy = []
        # First clear the previusly set data from rows
        for row in range(10, 32):
            self.ws.cell(row, 13).value = ""
            for col in range(2, 12):
                self.ws.cell(row, col).value = ""

        for idx, exp in enumerate(expenses):

            # Copy the selected file
            self.ws[self.exp_date_c + str(self.exp_start_r + idx)] = exp[0]
            self.ws[self.exp_desc_c + str(self.exp_start_r + idx)] = exp[1]
            self.ws[exp_type_table[exp[2]] + str(self.exp_start_r + idx)] = exp[4]
            self.ws[self.exp_curr_c + str(self.exp_start_r + idx)] = exp[3]
            self.ws[self.exp_exc_rate_c + str(self.exp_start_r + idx)] = exp[5]

            fts = {
                "path": exp[6].split("//")[1],
                "name": f"expense_[{idx + 1}]_{conf.now.month:02d}-{int(exp[0]):02d}_{exp[2]}.{exp[6].split('.')[1]}",
            }

            self.files_to_copy.append(fts)

        self.ready_to_save_signal.emit()

    @pyqtSlot(bool)
    def save(self, del_res: bool):
        if len(self.files_to_copy):
            self.set_user_data()
            self.wb.save(conf.get_expense_path())
            FolderHandler.copy_expenses(self.files_to_copy)
        FolderHandler.make_zip()
        if del_res:
            FolderHandler.delete_result_folder()
        self.saved_signal.emit()


class ExtraFilesSaver(QObject):
    def __init__(self, parent=None) -> None:
        super().__init__(parent=parent)

    @pyqtSlot(list)
    def copy_extra_files(self, extra_files: list):
        FolderHandler.copy_extra_files(extra_files)

def test():
    expenses = ExpensesDocGenerator()
    expenses.load()
    expenses.set_user_data()
    expenses.save()


if __name__ == "__main__":
    test()
